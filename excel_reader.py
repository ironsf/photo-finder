from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_products(
    path: Path,
    barcode_col: str = "A",
    name_col: str = "B",
    has_header: bool = False,
) -> list[tuple[str, str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    barcode_idx = column_index_from_string(barcode_col) - 1
    name_idx = column_index_from_string(name_col) - 1
    start_row = 2 if has_header else 1

    products = []
    for row in ws.iter_rows(min_row=start_row):
        code = _cell_to_str(row[barcode_idx].value) if barcode_idx < len(row) else ""
        if not code:
            continue
        name = _cell_to_str(row[name_idx].value) if name_idx < len(row) else ""
        products.append((code, name))

    wb.close()
    return products
